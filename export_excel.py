import sys
import json
import pandas as pd
import openpyxl
import re
from openpyxl.styles import Alignment, Font, Border, Side

# Regex to remove invalid XML 1.0 characters
# See https://www.w3.org/TR/REC-xml/#charsets
# Valid characters are:
# #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
_INVALID_XML_CHARS_RE = re.compile(
    u'[^\x09\x0A\x0D\x20-\xD7FF\xE000-\xFFFD\U00010000-\U0010FFFF]'
)

def clean_invalid_xml_chars(text):
    return _INVALID_XML_CHARS_RE.sub('', text) if isinstance(text, str) else text

def export_to_excel(data_json_string, output_file_path, column_fields_json_string):
    try:
        # JSON文字列をPythonのリスト/辞書に変換
        data = json.loads(data_json_string)
        column_fields = json.loads(column_fields_json_string)

        # DataFrameに変換
        # 'id'フィールドでヘッダー行を識別
        header_rows_data = [row for row in data if isinstance(row.get('id'), str) and row['id'].startswith('header_')]
        no_duty_row_data = [row for row in data if row.get('id') == 'row_no_duty']
        actual_data_rows = [row for row in data if isinstance(row.get('id'), int)]

        # Tabulatorの表示順序を維持するために、column_fieldsをそのまま使用
        # ただし、'id'フィールドはExcelには不要なので除外
        export_columns = [col for col in column_fields if col != 'id']

        # 各行データをDataFrameの形式に合わせる
        processed_data = []
        for row_data in data:
            processed_row = {}
            
            # Check if it's a header row
            is_header_row = isinstance(row_data.get('id'), str) and (row_data['id'].startswith('header_') or row_data['id'] == 'row_no_duty')

            # Populate processed_row with all export_columns, cleaning invalid XML chars
            for col_field in export_columns:
                value = row_data.get(col_field, '')
                
                # ヘッダー行の場合、'name'フィールドが空であればデフォルト値を設定（防御的な処理）
                if is_header_row and col_field == 'name' and not value:
                    value = {
                        'header_date': '日付',
                        'header_day': '曜日',
                        'header_holiday': '祝日',
                        'header_noon_night': '昼夜',
                        'row_no_duty': '当直不要'
                    }.get(row_data['id'], '')
                processed_row[col_field] = clean_invalid_xml_chars(value)
            # すべての行（空白行を含む）をエクスポート対象とする
            processed_data.append(processed_row)

        final_df = pd.DataFrame(processed_data, columns=export_columns)

        # 'duty_count'と'name'カラムのタイトルを調整
        # Tabulatorのヘッダー行はデータとして含まれているため、DataFrameのヘッダーは不要
        # ただし、'duty_count'と'name'はTabulatorのヘッダー行では空なので、
        # 実際のデータ行の最初のカラムとして扱うために、ここで調整は不要。
        # to_excel(header=False) で出力し、Tabulatorのヘッダー行がそのままExcelの最初の数行になる。
        
        # ExcelWriterを使用してExcelファイルに書き込み
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False) # header=FalseでDataFrameのヘッダーを出力しない
            workbook = writer.book
            sheet = writer.sheets['Sheet1']

            # 列幅の調整 (Tabulatorの幅設定を参考に)
            # column_fieldsの最初の2つがduty_countとnameであることを前提
            if 'duty_count' in export_columns:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(export_columns.index('duty_count') + 1)].width = 15
            if 'name' in export_columns:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(export_columns.index('name') + 1)].width = 15

            # その他の動的なカラム
            for i in range(len(export_columns)):
                if export_columns[i] != 'duty_count' and export_columns[i] != 'name':
                    col_letter = openpyxl.utils.get_column_letter(i + 1)
                    sheet.column_dimensions[col_letter].width = 4
            
            # 全てのセルを中央揃え
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"データを '{output_file_path}' にエクスポートしました。")
    except Exception as e:
        print(f"エラーが発生しました: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) > 2: # output_file_path と column_fields_json_string の2つ
        output_file_path = sys.argv[1]
        column_fields_json_string = sys.argv[2]
        data_json_string = sys.stdin.read() # 標準入力からJSONデータを読み込む
        export_to_excel(data_json_string, output_file_path, column_fields_json_string)
    else:
        print("使用法: python export_excel.py <output_file_path> <column_fields_json_string>", file=sys.stderr)
        sys.exit(1)