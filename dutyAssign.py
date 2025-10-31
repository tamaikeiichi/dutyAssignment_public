
import openpyxl
import numpy as np
import sys
import io
import os
import logging
import traceback
from ortools.sat.python import cp_model
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import pandas as pd

# 標準出力と標準エラー出力のエンコーディングをUTF-8に設定
# これにより、Electron(Node.js)側で文字化けせずに日本語を正しく受け取れるようになります。
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

def setup_logging():
    """ログ設定を初期化し、ユーザーのドキュメントフォルダにログを保存する"""
    try:
        # ユーザーのドキュメントフォルダのパスを取得
        # os.path.expanduser("~") はホームディレクトリを指す (例: C:\Users\YourUser)
        documents_path = os.path.join(os.path.expanduser("~"), "Documents")
        log_dir = os.path.join(documents_path, "DutyAssignmentLogs")

        # ログ用のディレクトリが存在しない場合は作成
        os.makedirs(log_dir, exist_ok=True)

        log_file_path = os.path.join(log_dir, 'duty_assign.log')

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[logging.FileHandler(log_file_path, 'w', 'utf-8')] # 'w'で起動時にログをリセット
        )
        logging.info("ログ設定が正常に完了しました。ログはここに記録されます。")
    except Exception as e:
        # ログ設定自体が失敗した場合、標準エラーに出力する
        print(f"致命的エラー: ログ設定に失敗しました。Error: {e}", file=sys.stderr)
        sys.exit(1)

# --- 定数定義 ---
# マッピング
VALUE_MAPPING = {"×": 0, "〇": 2, "輪番": 3, "\u3000": 1, " ": 1}
SHIFT_TYPE_MAPPING = {"昼": 0, "夜": 1}

# Excelレイアウトマーカー
MARKER_PERSON_START = "start"
MARKER_PERSON_END = "end"
MARKER_DATE_START = "start"
MARKER_DATE_END = "end"
MARKER_DATE_PAST = "past"

# Excel列/行インデックス
COL_REQUIRED_SHIFTS = 0
COL_NAMES = 1
ROW_MARKERS = 0
ROW_WEEKDAY = 1
ROW_DATE = 2
ROW_SHIFT_TYPE = 3

def create_schedule(file_path):
    # この関数が呼ばれた時点でログ設定が完了していることを確認
    if not logging.getLogger().hasHandlers():
        setup_logging()
    logging.info(f"処理開始: {file_path}")
    # Excelファイルの読み込み。1番目のシートを読む
    input_df = pd.read_excel(file_path, sheet_name=0, header=None)
    logging.info("Excelファイルの読み込み完了")
    
    # 勤務希望を数字に置き換え（〇：第一希望、空白：第二希望、×：勤務不可、輪番：当院管理当直とは無関係に必要な当直）
    # value_mapping = {"×": 0, "〇": 2, "輪番": 3, "\u3000": 1, " ": 1} # 定数に置き換え
    
    # 名前データの範囲を特定（start, endを元ファイルに書いておく）
    start_row = None 
    end_row = None 
    # 2列目のデータから名前の開始と終了行を探す
    for idx, val in enumerate(input_df.iloc[:, COL_NAMES]):
        if val == MARKER_PERSON_START:
            start_row = idx + 1  # 名前の開始行は "start" の次の行
        if val == MARKER_PERSON_END:
            end_row = idx
    if start_row is None or end_row is None:
        raise ValueError("Excel内に名前の'start'または'end'マーカーが見つかりませんでした。")
    logging.info(f"名前の範囲を特定: start_row={start_row}, end_row={end_row}")
    
    # 勤務希望の範囲を1行目から取得（past（先月のデータ始まり）, start, endを元ファイルに書いておく）
    past_col = None
    start_col = None
    end_col = None
    # 1行目のデータから "start"  "end" "past" を探す
    for idx, val in enumerate(input_df.iloc[ROW_MARKERS, :]):
        if val == MARKER_DATE_START:
            start_col = idx  
        if val == MARKER_DATE_END:
            end_col = idx+1
        if val == MARKER_DATE_PAST:
            past_col = idx
    
    if start_col is None or end_col is None or past_col is None:
        raise ValueError("Excel内に'start'または'end'または'past'マーカーが見つかりませんでした。")
    logging.info(f"日付の範囲を特定: start_col={start_col}, end_col={end_col}, past_col={past_col}")
    
    # 名前リスト
    names = input_df.iloc[0:end_row, COL_NAMES].tolist()
    # 個別対応をしたいときに使う
    ozaki_row = names.index("尾崎泰")  # 特定の人物の行番号を取得
    
    # 3行目の数字が日にち（昼夜で同じ数字が連続している場合は1日分である）
    date_numbers = input_df.iloc[ROW_DATE, 0:end_col].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int).values.tolist()
    date_numbers_withLastMonth = input_df.iloc[ROW_DATE, 0:end_col].tolist()
    date_numbers_onlyLastMonth = input_df.iloc[ROW_DATE, 0:start_col].tolist()
    
    # 必須勤務回数（Excelの1列目）を取り込む
    required_shifts = input_df.iloc[0:end_row, COL_REQUIRED_SHIFTS].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int).values.tolist()
    
    # 勤務タイプ列（昼 or 夜）の列インデックスを設定
    is_night = pd.to_numeric(input_df.iloc[ROW_SHIFT_TYPE, 0:end_col].replace(SHIFT_TYPE_MAPPING), errors='coerce').fillna(1).astype(int).tolist()
    
    # 日ごとの列インデックスリストを作成
    day_indices = []
    prev_num = None
    for idx, num in enumerate(date_numbers):
        if num != prev_num:
            day_indices.append([idx])
        else:
            day_indices[-1].append(idx)
        prev_num = num
    
    # 列インデックスから日付インデックス（ペアのインデックス）へのマッピングを作成。前月データを含まない。
    column_to_day_map = {
        col_idx: day_idx
        for day_idx, day_group in enumerate(day_indices)
        for col_idx in day_group
    }
    
    # 辞書のすべての値をリストに変換
    all_values = list(column_to_day_map.values())
    
    # リストの最後の要素を取得
    last_value = all_values[-1]
    
    # 前月データを含む全体のデータの〇×を数字に変換
    df_subset = input_df.iloc[0:end_row]
    df_numeric = df_subset.replace(
        VALUE_MAPPING
    ).fillna(1).apply(pd.to_numeric, errors='ignore')
    
    # debug:  df_numericをCSVファイルに書き出す
    # df_numeric.to_csv('df_numeric.csv', encoding='shift_jis', index=False)
    
    # 勤務希望の部分だけ取り出し（×や〇の部分）
    df_numeric.index = names
    
    # 日数（列数）と人数（行数）
    num_days = df_numeric.shape[1]
    num_people = df_numeric.shape[0]
    logging.info(f"データ準備完了: 人数={num_people}, 日数={num_days}")

    # --- 「〇」希望を出した人数の計算 ---
    # 目的の範囲をスライス
    target_range = df_numeric.iloc[start_row:end_row, start_col:end_col]
    # 各行に2が1つ以上含まれるか（any(axis=1)）を判定し、
    # Trueの行を合計（sum()）する
    num_rows_with_2 = (target_range == 2).any(axis=1).sum()
    logging.info(f"「〇」の希望を1つ以上出している人数: {num_rows_with_2}人")
    # --- ここまで ---
    
    # モデルの作成
    model = cp_model.CpModel()
    
    # 勤務変数の作成：x[i][d] = 1ならi番目のメンバーがd日目に勤務
    x = {}
    for i in range(start_row, end_row):
        for d in range(start_col, end_col):
            x[i, d] = model.NewBoolVar(f"x_{i}_{d}")
    logging.info("CP-SATモデルと変数の作成完了")
    
    # ========
    # 制約の設定 
    # ========
    
    # ×のところは割り当て不可
    for i in range(start_row, end_row):
        for d in range(start_col, end_col):
            if df_numeric.iloc[i, d] == 0:
                model.Add(x[i, d] == 0)
    
    # 勤務回数が指定回数に一致（ただし、輪番（= 3）は除外してカウント）
    for i in range(start_row, end_row):
        model.Add(
            sum(
                x[i, d] for d in range(start_col, end_col)
                if df_numeric.iloc[i, d] != 3
            ) == required_shifts[i]
        )
    
    # 各勤務（各列）につき、輪番以外で必ず1人割り当てる（輪番がいてもいなくても1人必要）
    for d in range(start_col, end_col):
        # 木曜日(外部から医師が派遣されるため、当メンバーでの当直不要)以外、割り当て人数が1人であるという制約を追加
        if input_df.iloc[ROW_WEEKDAY, d] != '木':
            model.Add(sum(x[i, d] 
                          for i in range(start_row, end_row)
                            if df_numeric.iloc[i, d] != 3) == 1
                            )
    
    # 輪番は記載通り割り当て（1日に複数の輪番でもOK）
    for i in range(start_row, end_row):
        for d in range(start_col, end_col):
            if df_numeric.iloc[i, d] == 3:
                model.Add(x[i, d] == 1)
    
    # 同じ人が7日未満に複数回勤務しないようにする制約 ## （夜勤務同士）
    for i in range(start_row, end_row):
        for d1 in range(start_col, end_col): # 最後の1日は調べない
            if is_night[d1] == 1:  # 夜勤務のみ制約を適用
                d1_position = column_to_day_map[d1]
                if d1_position != last_value: # 最後の日は調べない
                    if d1_position +6 < len(day_indices):  # 日数の範囲内であることを確認
                        last_d2 = d1_position + 6
                    else:
                        last_d2 = len(day_indices) -1  # 日数の範囲を超えないようにする
                    for d2 in range(day_indices[d1_position + 1][0], day_indices[last_d2][-1]+1):
                        # if d2 <= len(is_night):
                        if is_night[d2] == 1:  # 夜勤務同士のみ
                                model.Add(x[i, d1] + x[i, d2] <= 1)
    
    # 同じ人が7日未満に複数回勤務しないようにする制約 ## （昼->夜）
    for i in range(start_row, end_row):
        for d1 in range(start_col, end_col):
            if is_night[d1] != 1:  # 昼勤務
                d1_position = column_to_day_map[d1]
                if d1_position != last_value: # 最後の日は調べない
                    if d1_position +6 < len(day_indices):  # 日数の範囲内であることを確認
                        last_d2 = d1_position + 6
                    else:
                        last_d2 = len(day_indices) -1  # 日数の範囲を超えないようにする
                    for d2 in range(day_indices[d1_position + 1][0], day_indices[last_d2][-1]+1):
                        # if d2 <= len(is_night):
                        if is_night[d2] == 1:  # 夜勤務
                                model.Add(x[i, d1] + x[i, d2] <= 1)
    
    # 同じ人が7日未満に複数回勤務しないようにする制約 ## （夜->昼）
    for i in range(start_row, end_row):
        for d1 in range(start_col, end_col):
            if is_night[d1] == 1:  # 夜勤務
                d1_position = column_to_day_map[d1]
                if d1_position != last_value: # 最後の日は調べない
                    if d1_position +6 < len(day_indices):  # 日数の範囲内であることを確認
                        last_d2 = d1_position + 6
                    else:
                        last_d2 = len(day_indices) -1  # 日数の範囲を超えないようにする
                    for d2 in range(day_indices[d1_position + 1][0], day_indices[last_d2][-1]+1):
                        # if d2 <= len(is_night):
                        if is_night[d2] != 1:  # 昼勤務
                                model.Add(x[i, d1] + x[i, d2] <= 1)
    
    # 前月データに関して、同じ人が7日未満に複数回勤務しないようにする制約（夜勤務の場合。昼勤務を希望している場合は無視する）
    for i in range(start_row, end_row):
        for d1 in range(start_col, end_col):
            if is_night[d1] == 1:  # 夜勤務のみ制約を適用
                d1_position = column_to_day_map[d1] # 月頭から何日目かを示す
                start_d1_position = column_to_day_map[start_col]
                end_d1_position = start_d1_position + 5 # 月頭から6日目まで調べればいい
                if d1_position <= end_d1_position:
                    for d2 in range(day_indices[d1_position - 6][0], # １週間ごとの勤務はOK。
                                    day_indices[d1_position - 1][-1]):
                        if d2 >= 0:  # 前月データの範囲内であることを確認
                            if df_numeric.iloc[i, d2] >= 2:  # 前月データで勤務がある場合
                                if df_numeric.iloc[i, d1] != 3:  # 輪番希望であれば無視する（前月から1週間未満でも仕方がない）
                                    model.Add(x[i, d1] == 0)
                        else:
                            print("前月データが7日分以上コピーされていないため、前月データの制約は適用されません。")
    
    # 夜勤務の翌日は昼勤務不可
    for i in range(start_row, end_row):
        for d in range(start_col, end_col - 1):
            if is_night[d] == 1:  # 夜勤務の場合
                if is_night[d + 1] == 0:  # 翌日は昼勤務の場合
                    model.Add(x[i, d] + x[i, d + 1] <= 1)  # 翌日は昼勤務不可
    
    # 昼勤務の翌日は夜勤務不可
    for i in range(start_row, end_row):
        for d in range(start_col, end_col - 1):
            if is_night[d] != 1:  # 昼勤務の場合
                if is_night[d + 1] == 1:  # 続いて夜勤務の場合
                    if i != ozaki_row:  # 尾崎先生は昼勤務の翌日も夜勤務可能
                        model.Add(x[i, d] + x[i, d + 1] <= 1)  # 翌日は昼勤務不可
    
    # 昼夜連続勤務は、希望していなければ不可
    for i in range(start_row, end_row):
        for d in range(start_col, end_col):
            if is_night[d] == 0:  # 昼勤務の場合
                if df_numeric.iloc[i, d] != 2 or df_numeric.iloc[i, d + 1] != 2: # 昼夜両方に丸がなければ
                    model.Add(x[i, d] + x[i, d + 1] <= 1)
    
    # 同じ人が6日未満に複数回勤務しないようにする制約（昼->昼）（平日を跨げば可としたい）
    for i in range(start_row, end_row):
        for d1 in range(start_col, end_col):
            if is_night[d1] != 1:  # 昼勤務のみ制約を適用
                d1_position = column_to_day_map[d1]
                if d1_position != last_value: # 最後の日は調べない
                    if d1_position +5 < len(day_indices):  # 日数の範囲内であることを確認
                        last_d2 = d1_position + 5
                    else:
                        last_d2 = len(day_indices) -1 # 日数の範囲を超えないようにする
                    # if d1_position +7 < len(day_indices):  # 日数の範囲内であることを確認
                    for d2 in range(
                        day_indices[d1_position + 1][0], day_indices[last_d2][-1]+1):
                        # if d2 < len(is_night):
                        if is_night[d2] != 1:  # 昼勤務同士のみ
                            model.Add(x[i, d1] + x[i, d2] <= 1)
    
    # 輪番の後も7日未満は勤務不可（夜勤務のみ不可、昼はOK）
    for i in range(start_row, end_row):
        for d in range(start_col, end_col):
            if df_numeric.iloc[i, d] == 3:  # 輪番の日
                d_position = column_to_day_map[d]
                if d1_position +6 < len(day_indices):  # 日数の範囲内であることを確認
                    for offset in range(
                        day_indices[d1_position + 1][-1], day_indices[d1_position + 6][-1]):
                        nd = d + offset
                        if 0 <= nd < num_days and nd != d:
                            if is_night[nd] == 1:  # 夜勤務のみ不可
                                model.Add(x[i, nd] == 0)
    
    # 割り当て日数をカウントするための変数のリストを定義
    assigned_days_per_person = [
        model.NewIntVar(0, end_col, f'assigned_days_{i}')
        for i in range(0, end_row)
    ]
    for i in range(start_row, end_row):
        model.Add(
            assigned_days_per_person[i] ==
            sum(x[i, d] for d in range(start_col, end_col
                                       ) if df_numeric.iloc[i, d] == 2)
        )
    
    # =========
    # 各人の〇が1回以上割り当てられたかどうかを確認する
    # =========
    
    # 1. 各人に対してブール変数を定義
    # この変数は、その人が1日以上割り当てられた場合に True となります
    is_assigned_at_all = [
        model.NewBoolVar(f'is_assigned_at_all_{i}')
        for i in range(0, end_row)
    ]
    
    # 2. 割り当て日数とブール変数をリンクさせる制約を追加
    for i in range(start_row, end_row):
        # もし assigned_days_per_person[i] > 0 なら、is_assigned_at_all[i] は True
        model.Add(assigned_days_per_person[i] > 0
                  ).OnlyEnforceIf(is_assigned_at_all[i])
        # もし assigned_days_per_person[i] == 0 なら、is_assigned_at_all[i] は False
        model.Add(assigned_days_per_person[i] == 0
                  ).OnlyEnforceIf(is_assigned_at_all[i].Not())
    
    # 3. ブール変数の合計を計算
    # sum()はブール変数を自動的に 0 と 1 に変換して合計します
    # `is_assigned_at_all`リストの要素のうち、start_rowからend_rowの範囲だけを合計
    varianceOfAppliedMaru = sum(is_assigned_at_all[start_row:end_row])
    
    # 丸の総数を最大化
    # 〇が採用されれば2増える。空白が採用されれば1増える。
    # 3は輪番なので無視する
    totalAppliedMaru = sum(
        df_numeric.iloc[i, d] * x[i, d]
        for i in range(start_row, end_row)
        for d in range(start_col, end_col)
        if df_numeric.iloc[i, d] != 3
    )
    
    # 重み付け
    # 丸の採用数は下二けた
    weight_totalAppliedMaru = 1
    # 一つ以上採用された人数は上二けた
    # 〇を一人一つ以上採用することに重きを置いたscore
    weight_varianceOfAppliedMaru = 1000
    
    # 最終的な目的関数
    model.Maximize(totalAppliedMaru * weight_totalAppliedMaru +
                    varianceOfAppliedMaru * weight_varianceOfAppliedMaru)
    logging.info("全ての制約と目的関数の設定完了")
    
    # --- ソルバーの実行とログのファイル保存 ---
    solver = cp_model.CpSolver()
    # ソルバーの探索ログを有効にする
    # solver.parameters.log_search_progress = True
    # CP-SATソルバーの場合
    # solver.parameters.num_search_workers = 1  # 1スレッドに限定
    # solver.parameters.max_time_in_seconds = 1.0 # 1秒でテスト
    solver.parameters.use_lns_only = True  # LNSのみを使用
    # 実行不可能な場合(INFEASIBLE)に、原因となっている制約の調査を有効にする
    # solver.parameters.num_workers = 1
    # solver.parameters.log_infeasible_subsystem = True
    # ソルバーのログを保存するファイルパスを定義
    documents_path = os.path.join(os.path.expanduser("~"), "Documents")
    log_dir = os.path.join(documents_path, "DutyAssignmentLogs")
    solver_log_path = os.path.join(log_dir, 'solver_log.log')
    # 元の標準出力を保存
    original_stdout = sys.stdout
    logging.info("ソルバーの実行開始")
    # try:
    #     # 標準出力をファイルにリダイレクト
    #     with open(solver_log_path, 'w', encoding='utf-8') as f:
    #         print("実行開始５", flush=True)
    #         sys.stdout = f
    status = solver.Solve(model)
    # finally:
    #     # 標準出力を元に戻す
    #     sys.stdout = original_stdout
    
    logging.info(f"ソルバーの実行完了. ステータス: {solver.StatusName(status)}")
    
    # 最適解が得られたか確認
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        # 最適解の目的関数の値を取得
        optimal_score = solver.ObjectiveValue()

        # スコアを「〇が1つ以上割り当てられた人数(上位桁)」と「希望点の合計(下位桁)」に分割
        # 上位桁: 〇が1つ以上割り当てられた人数
        num_people_assigned_score = int(optimal_score // weight_varianceOfAppliedMaru)
        # 下位桁: 希望点の合計
        total_preference_score = int(optimal_score % weight_varianceOfAppliedMaru)

        # f-string内で改行文字 \n を使うことで、出力を複数行に分割
        print(f"最大スコア: {optimal_score:.0f}\n"
              f"内訳:\n"
              f"  - 〇を一つ以上記載した人数: {num_rows_with_2}人\n"
              f"  - 〇を一つ以上採用した人数: {num_people_assigned_score}人\n"
              f"  - 〇の採用スコア: {total_preference_score}")
        # 最適解の各変数の値からresult_matrixを再構築
        result_matrix = []
        for i in range(start_row, end_row):
            row = []
            for d in range(start_col, end_col):
                if df_numeric.iloc[i, d] == 3:
                    val = 3
                else:
                    # ソルバーから変数の値を取得
                    val = solver.Value(x[i, d])
                row.append(val)
            result_matrix.append(row)
    
        # 元データを使用してIndex情報を作成
        header_data = input_df.iloc[0:4, start_col:end_col].copy()
        header_data.fillna('', inplace=True)
        header_array = header_data.values
        combined_matrix = np.vstack((header_array, result_matrix)) 
        # 列（日付）と行（名前）のラベル付きDataFrameを作成
        pd.set_option('future.no_silent_downcasting', True)
        result_df = pd.DataFrame(combined_matrix, index=names)
        # result_df = result_df.astype(object) # 全体をobject型（文字列）に変換しておく
        # 〇×表記＋輪番に変換
        result_df.iloc[range(start_row, end_row),] = (
        result_df.iloc[range(start_row, end_row),].replace({1: "〇", 0: "", 3: "輪番"})
    )
        display_df = result_df.copy()
    
        # 出力ファイル名を動的に生成
        # 入力ファイルと同じディレクトリに出力ファイルを作成
        output_dir = os.path.dirname(file_path)
        # スコアを小数点1位に丸めてファイル名に含める
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = os.path.join(output_dir, f"{base_filename}_assigned_score{optimal_score:.0f}.xlsx")
    
        logging.info(f"結果をExcelファイルに書き込み開始: {output_filename}")
        # 以降はExcelに見栄えよく書き出すための設定（元のコードのまま）
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='Sheet1', header=False)
            worksheet = writer.sheets['Sheet1']
    
            # 列幅の調整
            worksheet.column_dimensions['A'].width = 15
            for j in range(len(display_df.columns)):
                column_letter = openpyxl.utils.get_column_letter(j + 2)
                worksheet.column_dimensions[column_letter].width = 4
            
            # 全ての行の高さを指定
            for j in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[j].height = 20
    
            # 全てのセルのフォントサイズを12に設定
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(size=12)
    
            # 罫線と背景色の設定
            white_thick_border = Border(left=Side(style='thick', color='FFFFFF'),
                                        right=Side(style='thick', color='FFFFFF'),
                                        top=Side(style='thick', color='FFFFFF'),
                                        bottom=Side(style='thick', color='FFFFFF'))
            light_red_fill_light = PatternFill(
                start_color="FFe0e0", end_color="FFe0e0", fill_type="solid")
            light_red_fill_dark = PatternFill(
                start_color="FFd0d0", end_color="FFd0d0", fill_type="solid")
            light_yellow_fill_light = PatternFill(
                start_color="FAFAA0", end_color="FAFAA0", fill_type="solid")
            light_yellow_fill_dark = PatternFill(
                start_color="FAFA20", end_color="FAFA20", fill_type="solid")
            light_grey_fill_light = PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            light_grey_fill_dark = PatternFill(
                start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
            center_alignment = Alignment(
                horizontal='center', vertical='center')
    
            weekdays_header = display_df.columns.get_level_values(0)
            is_holiday_col = [
                any(day in str(wd) for day in [
                    "土", "日", "祝"
                    ]) for wd in weekdays_header]
            is_thirsday_col = [wd == "木" for wd in weekdays_header]
    
            row_of_weekdays = display_df.iloc[1]
    
            # その行をループし、「日」の列のインデックスを見つける
            holidays = ["土", "日", "祝"]
            holiday_column_indices = [
                i for i, val in enumerate(row_of_weekdays) if val in holidays]
    
            weekdays = ["月", "火", "水", "金"]
            weekdays_header_indices = [
                i for i, val in enumerate(row_of_weekdays) if val in weekdays]
    
            thirsdays = ["木"]
            thirsday_header_indices = [
                i for i, val in enumerate(row_of_weekdays) if val in thirsdays]
    
            for col in holiday_column_indices:
                alternate_row_color = False
                for row in range(2, end_row + 1):
                    alternate_row_color = not alternate_row_color
                    cell = worksheet.cell(row=row, column=col + 2) # +2は名前列と1行目のヘッダーをスキップ
                    cell.fill = light_red_fill_light if alternate_row_color else light_red_fill_dark
    
            for col in weekdays_header_indices:
                alternate_row_color = False
                for row in range(2, end_row + 1):
                    alternate_row_color = not alternate_row_color
                    cell = worksheet.cell(row=row, column=col + 2) # +2は名前列と1行目のヘッダーをスキップ
                    cell.fill = light_yellow_fill_light if alternate_row_color else light_yellow_fill_dark
            
            for col in thirsday_header_indices:
                alternate_row_color = False
                for row in range(2, end_row + 1):
                    alternate_row_color = not alternate_row_color
                    cell = worksheet.cell(row=row, column=col + 2) # +2は名前列と1行目のヘッダーをスキップ
                    cell.fill = light_grey_fill_light if alternate_row_color else light_grey_fill_dark
    
            col = 1 # 名前列
            alternate_row_color = False
            for row in range(2, end_row + 1):   
                alternate_row_color = not alternate_row_color
                cell = worksheet.cell(row=row, column=col) # 名前列
                cell.fill = light_grey_fill_light if alternate_row_color else light_grey_fill_dark
    
            for row in range(1, end_row + 1):
                for col in range(1, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    # 罫線を適用
                    cell.border = white_thick_border
            
            # ヘッダーとインデックスのフォントを通常書体に変更
            num_header_rows = display_df.columns.nlevels
            for row_idx in range(1, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    if row_idx <= num_header_rows or col_idx == 1:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.font and cell.font.bold:
                            cell.font = cell.font.copy(bold=False)
    
            # シートの全てのセルをループ処理
            for row in range(1, end_row + 1):
                for col in range(1, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = center_alignment
        
        logging.info("Excelファイルへの書き込み完了")
        # print(f"勤務表を'{output_filename}'に保存しました。")
        return f"勤務表を'{output_filename}'に保存しました。"
    
    else:
        # print("最適解が見つかりませんでした。")
        logging.warning("最適解が見つかりませんでした。")
        return "最適解が見つかりませんでした。"

if __name__ == "__main__":
    # スクリプト実行の最初にログ設定を呼び出す
    setup_logging()

    try:
        # コマンドライン引数からファイルパスを取得
        if len(sys.argv) > 1:
            file_path_arg = sys.argv[1]
            # 処理を実行して結果を標準出力に出力
            result_message = create_schedule(file_path_arg)
            print(result_message)
        else:
            logging.error("コマンドライン引数にファイルパスが指定されていません。")
            print("Error: No file path provided.")
            sys.exit(1)
    except Exception as e:
        # 予期せぬエラーをすべてキャッチしてログに記録
        logging.critical(f"スクリプト実行中に致命的なエラーが発生しました: {e}")
        logging.critical(traceback.format_exc()) # スタックトレースをログに出力
        print(f"An unexpected error occurred: {e}", file=sys.stderr)
        sys.exit(1)